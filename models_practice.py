from sqlalchemy import create_engine, MetaData, insert, select, text
from datetime import datetime
import openpyxl

host = "localhost"
port = 5432
user = "psychoslvt"
password = "Bghujknmol123"
db = "demo"

engine = create_engine(f"postgresql+psycopg2://{user}:{password}@{host}:{port}/{db}")

metadata = MetaData()
metadata.reflect(bind=engine)

User = metadata.tables["User"]
Delivery = metadata.tables["Delivery"]
Order = metadata.tables["Order"]
OrderItem = metadata.tables["OrderItem"]
Product = metadata.tables["Product"]


# def parse_date(value):
#     s = str(value).split()[0]
#     isinstance(value, datetime)
#     for fmt in ("%Y-%m-%d", "%m/%d/%Y"):
#         try:
#             return datetime.strptime(s, fmt).date()
#         except:
#             pass
#     raise ValueError(f"Unknown date format: {value}")

def normalize_article(data: str):
    raw = [x.strip() for x in data.split(',')]
    articles = []
    quantities = []

    for i in range(0, len(raw), 2):
        articles.append(raw[i])
        quantities.append(int(raw[i+1]))

    return articles, quantities


if __name__ == "__main__":
    with engine.begin() as conn:
        conn.execute(text('TRUNCATE "OrderItem" CASCADE; TRUNCATE "Order" CASCADE; TRUNCATE "Product" CASCADE; TRUNCATE "User" CASCADE; TRUNCATE "Delivery" CASCADE;'))
    
    user_table = openpyxl.load_workbook("import/user_import.xlsx").active

    user_rows = []
    for row in user_table.iter_rows(values_only=True, min_row=2):
        if row[0] is None:
            continue
        data = {
            "role": str(row[0]),
            "name": str(row[1]),
            "login": str(row[2]),
            "password": str(row[3])
        }

        user_rows.append(data)

    with engine.begin() as conn:
        conn.execute(insert(User), user_rows)

    with engine.connect() as conn:
        user_map = dict(conn.execute(select(User.c.name, User.c.id).where(User.c.role != "Авторизованный клиент")).fetchall())

    delivery_table = openpyxl.load_workbook("import/Пункты выдачи_import.xlsx").active

    delivery_rows = []
    for i, row in enumerate(delivery_table.iter_rows(values_only=True)):
        if row[0] is None:
            continue
        data = {
            "id": i+1,
            "address": str(row[0])
        }

        delivery_rows.append(data)

    with engine.begin() as conn:
        conn.execute(insert(Delivery), delivery_rows)

    product_table = openpyxl.load_workbook("import/Tovar.xlsx").active

    product_rows = []

    for row in product_table.iter_rows(values_only=True, min_row=2):
        if row[0] is None:
            continue
        data = {
            "article": str(row[0]),
            "title": str(row[1]),
            "measure_type": str(row[2]),
            "price": float(row[3]),
            "supplier": str(row[4]),
            "producer": str(row[5]),
            "category": str(row[6]),
            "discount": float(row[7]),
            "quantity": int(row[8]),
            "description": str(row[9]),
            "image_url": f"import/{str(row[10])}" if row[10] is not None else None
        }

        product_rows.append(data)

    with engine.begin() as conn:
        conn.execute(insert(Product), product_rows)

    order_table = openpyxl.load_workbook("import/Заказ_import.xlsx").active

    orders_rows = []
    order_item_rows = []

    for row in order_table.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue

        try:
            order_id = int(row[0])
            article_raw = str(row[1])
            order_date = row[2]
            delivery_date = str(row[3])
            if not isinstance(order_date, datetime):
                continue
            delivery_id = int(row[4])
            user_id = user_map.get(str(row[5]))
            challenge_code = int(row[6])
            status = str(row[7])
        except:
            continue

        orders_rows.append({
            "id": order_id,
            "order_date": order_date,
            "delivery_date": delivery_date,
            "address_id": delivery_id,
            "user_id": user_id,
            "challenge_code": challenge_code,
            "status": status
        })

        articles, quantities = normalize_article(article_raw)
        for a, q in zip(articles, quantities):
            order_item_rows.append({
                "article": a,
                "order": order_id,
                "quantity": q
            })
        
    with engine.begin() as conn:
        conn.execute(insert(Order), orders_rows)
        conn.execute(insert(OrderItem), order_item_rows)
        

    