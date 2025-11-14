import base64
import hashlib
import random
from datetime import datetime
from enum import StrEnum
import openpyxl

from sqlalchemy import Integer, Column, ForeignKey, String, Float, Date, UniqueConstraint, text, CheckConstraint
from sqlalchemy.orm import Mapped, mapped_column, relationship

from db.base import Base
from db.connection import engine, create_session
from settings import app_settings
from settings import app_settings


def f(string: str):
            string = string.split()[0]
            try:
                return datetime.strptime(string, "%Y-%m-%d")
            except Exception:
                return datetime.strptime(string, "%m/%d/%Y")

class User(Base):
    __tablename__ = "User"

    __table_args__ = (UniqueConstraint('name', 'role'),)

    id: Mapped[int] = mapped_column(primary_key=True)
    role: Mapped[str] = mapped_column(String(30))
    name: Mapped[str] = mapped_column(String(50))
    login: Mapped[str] = mapped_column(String(50))
    password: Mapped[str] = mapped_column(String(50))
    orders = relationship("Order", back_populates="user", uselist=True)

class Product(Base):
    __tablename__ = "Product"

    id: Mapped[int] = mapped_column(primary_key=True)
    title: Mapped[str] = mapped_column(String(30))
    measure_type: Mapped[str] = mapped_column(String(30))
    price: Mapped[float] = mapped_column(Float)
    supplier: Mapped[str] = mapped_column(String(30))
    producer: Mapped[str] = mapped_column(String(30))
    category: Mapped[str] = mapped_column(String(30))
    discount: Mapped[float] = mapped_column(Float)
    quantity: Mapped[int] = mapped_column()
    description: Mapped[str] = mapped_column(String(1024))
    image_url: Mapped[str] = mapped_column(String(100))

class Delivery(Base):
    __tablename__ = "Delivery"

    id: Mapped[int] = mapped_column(primary_key=True)
    address: Mapped[str] = mapped_column(String(256))

class Order(Base):
    __tablename__ = "Order"

    id: Mapped[int] = mapped_column(primary_key=True)
    article: Mapped[str] = mapped_column(String(50))
    order_date: Mapped[Date] = mapped_column(Date)
    delivery_date: Mapped[Date] = mapped_column(Date)
    address_id: Mapped[int] = mapped_column(ForeignKey("Delivery.id"))
    user_id: Mapped[int] = mapped_column(ForeignKey("User.id"), nullable=True)
    user = relationship(User, back_populates="orders")
    challenge_code: Mapped[int] = mapped_column()
    status: Mapped[str] = mapped_column(String(30))


if __name__ == '__main__':
    with engine.connect() as conn:
        for table in Base.metadata.tables:
            conn.execute(text(f"DROP TABLE IF EXISTS \"{table}\" CASCADE"))
        conn.commit()
    Base.metadata.create_all(engine)

    with create_session() as session:
        user_table = openpyxl.load_workbook("db/import/user_import.xlsx")

        user_sheet = user_table.active

        user_data = {
            str(row[1].value): User(
                role=str(row[0].value),
                name=str(row[1].value),
                login=str(row[2].value),
                password=str(row[3].value),
            )
            for row in list(user_sheet.iter_rows())[1:] if row[0].value is not None
        }
        session.add_all(user_data.values())

        delivery_table = openpyxl.load_workbook("db/import/Пункты выдачи_import.xlsx")

        delivery_sheet = delivery_table.active

        delivery_data = [
            Delivery(
                id=i + 1,
                address = str(row[0].value)
            )
            for i, row in enumerate(list(delivery_sheet.iter_rows())[1:])  if row[0].value is not None
        ]
        session.add_all(delivery_data)


        order_table = openpyxl.load_workbook("db/import/Заказ_import.xlsx")

        order_sheet = order_table.active
        filtered_user_data = {k: v for k, v in user_data.items() if v.role == "Авторизированный клиент"}

        order_data = [
            Order(
                id=int(row[0].value),
                article=str(row[1].value),
                order_date=f(str(row[2].value)),
                delivery_date=str(row[3].value),
                address_id=int(row[4].value),
                user=filtered_user_data.get(str(row[5].value)),
                challenge_code=int(row[6].value),
                status=str(row[7].value)
            )
            for row in list(order_sheet.iter_rows())[1:]  if row[0].value is not None
        ]
        session.add_all(order_data)



        # products_table = openpyxl.load_workbook("db/import/Tovar.xlsx")

        # products_sheet = products_table.active

        # products_data = [
        #     Product(
        #         article=str(row[0].value),
        #         title=str(row[1].value),
        #         measure_type=str(row[2].value),
        #         price=float(row[3].value),
        #         supplier=str(row[4].value),
        #         producer=str(row[5].value),
        #         category=str(row[6].value),
        #         discount=float(row[7].value),
        #         quantity=int(row[8].value),
        #         description=str(row[9].value),
        #         image_url=f"{app_settings.root}/assets/images/{str(row[10])}",
        #     )
        #     for row in list(products_sheet.iter_rows())[1:] if row[0].value is not None
        # ]
        # session.add_all(products_data)
