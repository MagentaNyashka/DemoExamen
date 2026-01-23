import openpyxl
from datetime import datetime
from sqlalchemy import create_engine, MetaData, insert, text


host = "localhost"
port = 5432
db = "demo"
user = "psychoslvt"
password = "Bghujknmol123"
echo = False

engine = create_engine(f"postgresql+psycopg2://{user}:{password}@{host}:{port}/{db}", echo=echo)


def parse_date(value):
    if value is None:
        return None
    s = str(value).split()[0]
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except:
            pass
    raise ValueError(f"Unknown date format: {value}")


def normalize_article(data: str):
    raw = [x.strip() for x in data.split(',')]
    articles = []
    quantities = []

    for i in range(0, len(raw), 2):
        articles.append(raw[i])
        quantities.append(int(raw[i+1]))

    return articles, quantities


# -----------------------------
#      REFLECT TABLES
# -----------------------------
metadata = MetaData()
metadata.reflect(bind=engine)

User = metadata.tables["User"]
Delivery = metadata.tables["Delivery"]
Product = metadata.tables["Product"]
Order = metadata.tables["Order"]
OrderItem = metadata.tables["OrderItem"]


if __name__ == "__main__":
    # -----------------------------
    #       TRUNCATE TABLES
    # -----------------------------
    with engine.begin() as conn:
        conn.execute(text('TRUNCATE "OrderItem" CASCADE; TRUNCATE "Order" CASCADE; TRUNCATE "Product" CASCADE; TRUNCATE "User" CASCADE; TRUNCATE "Delivery" CASCADE'))


    # -----------------------------
    #       IMPORT USERS
    # -----------------------------
    user_table = openpyxl.load_workbook("import/user_import.xlsx").active

    user_rows = [
        {
            "role": str(row[0].value),
            "name": str(row[1].value),
            "login": str(row[2].value),
            "password": str(row[3].value),
        }
        for row in user_table.iter_rows(min_row=2)
        if row[0].value is not None
    ]

    with engine.begin() as conn:
        conn.execute(insert(User), user_rows)

    # Кэш для связи заказ → user_id
    with engine.connect() as conn:
        user_map = dict(conn.execute(select(User.c.name, User.id).where(User.c.role != "Авторизованный клиент")).fetchall())
    #    user_map = dict(conn.execute(text("SELECT name, id FROM \"User\" WHERE role <> 'Авторизированный клиент';")).fetchall())
        # user_map = dict(conn.execute(text("SELECT name, id FROM \"User\";")).fetchall())



    # -----------------------------
    #       IMPORT DELIVERY
    # -----------------------------
    delivery_table = openpyxl.load_workbook("import/Пункты выдачи_import.xlsx").active

    delivery_rows = [
        {"id": i + 1, "address": str(row[0].value)}
        for i, row in enumerate(delivery_table.iter_rows(min_row=2))
        if row[0].value is not None
    ]

    with engine.begin() as conn:
        conn.execute(insert(Delivery), delivery_rows)

    # -----------------------------
    #       IMPORT PRODUCTS
    # -----------------------------
    products_table = openpyxl.load_workbook("import/Tovar.xlsx").active

    products_rows = [
        {
            "article": str(row[0].value),
            "title": str(row[1].value),
            "measure_type": str(row[2].value),
            "price": float(row[3].value),
            "supplier": str(row[4].value),
            "producer": str(row[5].value),
            "category": str(row[6].value),
            "discount": float(row[7].value),
            "quantity": int(row[8].value),
            "description": str(row[9].value),
            "image_url": f"import/{row[10].value}" if row[10].value is not None else "import/picture.png",
        }
        for row in products_table.iter_rows(min_row=2)
        if row[0].value is not None
    ]

    with engine.begin() as conn:
        conn.execute(insert(Product), products_rows)

    # -----------------------------
    #         IMPORT ORDERS
    # -----------------------------
    order_table = openpyxl.load_workbook("import/Заказ_import.xlsx").active

    orders_rows = []
    order_item_rows = []

    for row in order_table.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue

        try:
            order_id = int(row[0])
            article_raw = str(row[1])
            order_date = parse_date(row[2])
            delivery_date = parse_date(row[3])
            address_id = int(row[4])
            user_name = str(row[5])
            user_id = user_map.get(user_name)
            challenge = int(row[6])
            status = str(row[7])
        except:
            continue

        # создаём запись заказа
        orders_rows.append({
            "id": order_id,
            "order_date": order_date,
            "delivery_date": delivery_date,
            "address_id": address_id,
            "user_id": user_id,
            "challenge_code": challenge,
            "status": status
        })

        # создаём позиции заказа
        articles, quantities = normalize_article(article_raw)
        for a, q in zip(articles, quantities):
            order_item_rows.append({
                "order": order_id,
                "article": a,
                "quantity": q
            })

    with engine.begin() as conn:
        conn.execute(insert(Order), orders_rows)
        conn.execute(insert(OrderItem), order_item_rows)
